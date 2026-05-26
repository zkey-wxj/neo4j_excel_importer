"""graph_export.py — 图谱数据导出为 Excel / JSON。"""
from __future__ import annotations

import io
import json
import logging
from typing import Any

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

logging.getLogger("openpyxl").setLevel(logging.ERROR)


def nodes_to_rows(nodes: list[dict[str, Any]]) -> tuple[list[str], list[list[Any]]]:
    """将节点列表转为 (headers, rows)，适用于 Excel 和 CSV 导出。"""
    headers = ["nid", "name", "labels", "description"]
    extra_keys: set[str] = set()
    for n in nodes:
        extra_keys.update(n.get("properties", {}).keys())
    headers.extend(sorted(extra_keys))

    rows: list[list[Any]] = []
    for n in nodes:
        row: list[Any] = [
            n.get("nid", ""),
            n.get("name", ""),
            ", ".join(n.get("labels", [])),
            n.get("description") or "",
        ]
        props = n.get("properties", {})
        for h in headers[4:]:
            row.append(props.get(h, ""))
        rows.append(row)
    return headers, rows


def relations_to_rows(relations: list[dict[str, Any]]) -> tuple[list[str], list[list[Any]]]:
    """将关系列表转为 (headers, rows)，适用于 Excel 和 CSV 导出。"""
    headers = ["source_nid", "rel_type", "target_nid", "description"]
    extra_keys: set[str] = set()
    for r in relations:
        extra_keys.update(r.get("properties", {}).keys())
    headers.extend(sorted(extra_keys))

    rows: list[list[Any]] = []
    for r in relations:
        row: list[Any] = [
            r.get("source_nid", ""),
            r.get("rel_type", ""),
            r.get("target_nid", ""),
            r.get("description") or "",
        ]
        props = r.get("properties", {})
        for h in headers[4:]:
            row.append(props.get(h, ""))
        rows.append(row)
    return headers, rows


def export_excel(nodes: list[dict[str, Any]], relations: list[dict[str, Any]], sheet_name: str = "知识图谱") -> bytes:
    """将节点和关系导出到单个 xlsx sheet，中间间隔 2 个空行。"""
    wb = Workbook()
    ws: Worksheet = wb.create_sheet(sheet_name, 0)  # type: ignore[assignment]
    del wb[wb.sheetnames[1]]

    node_headers, node_rows = nodes_to_rows(nodes)
    ws.append(node_headers)
    for row in node_rows:
        ws.append(row)

    ws.append([])
    ws.append([])

    rel_headers, rel_rows = relations_to_rows(relations)
    ws.append(rel_headers)
    for row in rel_rows:
        ws.append(row)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_json(nodes: list[dict[str, Any]], relations: list[dict[str, Any]]) -> bytes:
    """将节点和关系导出为 JSON 字节流。"""
    data = {"nodes": nodes, "relations": relations}
    return json.dumps(data, indent=2, ensure_ascii=False).encode("utf-8")
